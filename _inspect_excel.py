# -*- coding: utf-8 -*-
import openpyxl
p = r'C:\Users\15831\Desktop\能力验证平台验收\测试用例-AI版.xlsx'
wb = openpyxl.load_workbook(p)
print('sheets:', wb.sheetnames)
for ws in wb.worksheets:
    print('---', ws.title, 'max_row:', ws.max_row, 'max_col:', ws.max_column)
    print('hdr:', [c.value for c in ws[1]])
    for r in (2, 3):
        vals = [str(c.value)[:25] if c.value is not None else None for c in ws[r]]
        print('row%d:' % r, vals)
