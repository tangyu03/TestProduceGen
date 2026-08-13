# -*- coding: utf-8 -*-
"""把 PT017_output.md 同步到 测试用例-AI版.xlsx。

只覆盖来自原文的列: 用例标识 / 模块 / 用例名称 / 对应需求条目 / 前置条件(Given) / 测试步骤(When) / 预期结果(Then)。
保留第1行表头和第2行示例行; 其余列(测试类型/对应需求内容/测试数据/实测结果/是否通过/备注/测试人员/测试日期)不动。

用法:
    python sync_md_to_ai_excel.py [输出.md] [目标.xlsx]
路径缺省时回落当前目录: MD='PT017_output.md', XLSX=第一个 '测试用例*.xlsx'。
"""
import re
import sys
from pathlib import Path
import openpyxl

# 输入/输出路径: 命令行参数优先, 缺省回落当前目录相对路径
MD = sys.argv[1] if len(sys.argv) > 1 else 'PT017_output.md'
if len(sys.argv) > 2:
    XLSX = sys.argv[2]
else:
    _cands = sorted(Path('.').glob('测试用例*.xlsx'))
    XLSX = str(_cands[0]) if _cands else '测试用例-AI版.xlsx'

md_path = Path(MD)
if not md_path.exists():
    print(f"ERROR: 未找到 Markdown 文件: {MD}")
    sys.exit(1)
if not Path(XLSX).exists():
    print(f"ERROR: 未找到 Excel 文件: {XLSX}")
    sys.exit(1)

text = md_path.read_text(encoding='utf-8')
blocks = re.split(r'\n(?=### )', text)

entries = []
for b in blocks:
    b = b.strip()
    if not b.startswith('### '):
        continue
    m = re.match(r'### (PROC-\d+)[：:]?\s*(.*)', b)
    if not m:
        continue
    pid = m.group(1)
    title = m.group(2).strip()
    cov = re.search(r'\*\*覆盖需求\*\*：(.+)', b)
    coverage = cov.group(1).strip() if cov else ''
    biz = re.search(r'\*\*业务定位\*\*：(.+)', b)
    biz_val = biz.group(1).strip() if biz else ''
    # 模块 = 最后一个 "｜" 段，剥离 phase_label 内嵌的落段后缀（P{相位}·{原因}）
    for sep in ('｜', '|'):
        if sep in biz_val:
            module = biz_val.split(sep)[-1].strip()
            break
    else:
        module = biz_val
    module = re.sub(r'（[^（）]*）$', '', module)
    gsec = re.search(r'\*\*Given\*\*(.*?)(?=\*\*|\Z)', b, re.S)
    given = gsec.group(1).strip() if gsec else ''
    wsec = re.search(r'\*\*When\*\*(.*?)(?=\*\*|\Z)', b, re.S)
    when = wsec.group(1).strip() if wsec else ''
    tsec = re.search(r'\*\*Then\*\*(.*?)(?=\*\*依赖\*\*|\*\*弱依赖\*\*|\Z)', b, re.S)
    then = tsec.group(1).strip() if tsec else ''
    entries.append((pid, title, module, coverage, given, when, then))

print(f'parsed {len(entries)} entries')

wb = openpyxl.load_workbook(XLSX)
ws = wb.active
print('sheet:', ws.title, 'max_row:', ws.max_row)

hdr = {str(c.value).strip(): c.column for c in ws[1] if c.value is not None}
print('headers:', hdr)
col_id = hdr['用例标识']
col_mod = hdr.get('模块')  # 可选列: 部分模板无此列, 缺省跳过
col_name = hdr['用例名称']
col_req = hdr['对应需求条目']
col_pre = hdr['前置条件']
col_step = hdr['测试步骤']
col_exp = hdr['预期结果']

# 写入列集合 (模块缺失时排除)
src_cols = [col_id, col_name, col_req, col_pre, col_step, col_exp]
if col_mod is not None:
    src_cols.append(col_mod)

START = 3  # 数据从第3行开始(第1行表头, 第2行示例)
n_existing = ws.max_row - START + 1
n_new = len(entries)

# 覆盖已有数据行
for i, (pid, title, module, coverage, given, when, then) in enumerate(entries):
    r = START + i
    ws.cell(row=r, column=col_id, value=pid)
    ws.cell(row=r, column=col_name, value=title)
    ws.cell(row=r, column=col_req, value=coverage)
    ws.cell(row=r, column=col_pre, value=given)
    ws.cell(row=r, column=col_step, value=when)
    ws.cell(row=r, column=col_exp, value=then)
    if col_mod is not None:
        ws.cell(row=r, column=col_mod, value=module)

# 若原文条目比文件现有行少, 清空多余行的来源列
for r in range(START + n_new, START + n_existing):
    for c in src_cols:
        ws.cell(row=r, column=c).value = None

wb.save(XLSX)
print(f'synced {n_new} rows (rows {START}..{START+n_new-1}); file total rows now {ws.max_row}')
