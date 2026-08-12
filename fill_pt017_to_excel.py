# -*- coding: utf-8 -*-
"""把 PT017_output.md 的测试规程填入 测试用例+记录-V1.1.xlsx。

映射(按用户要求):
  用例标识    A = PROC-XXX
  模块        B = 业务定位中 "｜" 后面的部分 (新增列, 在用例标识后面)
  用例名称    C = 标题去掉 "PROC-XXX：" 的部分
  测试类型    D = 留空
  对应需求条目 E = 覆盖需求
  对应需求内容 F = 留空
  前置条件    G = given 段
  测试步骤    H = when 段
  测试数据    I = 不填
  预期结果    J = then 段
其余列(实测结果/是否通过/备注/测试人员/测试日期)不填。

用法: python fill_pt017_to_excel.py [目标xlsx路径]
  - 不传参数时用默认路径。
  - 目标文件应是从模板恢复(只有表头+示例行)的状态; 若已填充会从 max_row+1 追加。
"""
import re
import sys
import openpyxl

MD = r'C:\Users\15831\Downloads\final_pkgV\PT017_output.md'
XLSX = r'C:\Users\15831\Desktop\航天软件功能性\测试用例+记录-V1.1.xlsx'

if len(sys.argv) > 1:
    XLSX = sys.argv[1]

text = open(MD, encoding='utf-8').read()
blocks = re.split(r'\n(?=### )', text)

entries = []
for b in blocks:
    b = b.strip()
    if not b.startswith('### '):
        continue
    m = re.match(r'### (PROC-\d+)[：:]?\s*(.*)', b)
    if not m:
        continue
    pid = m.group(1)
    title = m.group(2).strip()
    cov = re.search(r'\*\*覆盖需求\*\*：(.+)', b)
    coverage = cov.group(1).strip() if cov else ''
    biz = re.search(r'\*\*业务定位\*\*：(.+)', b)
    biz_val = biz.group(1).strip() if biz else ''
    # 模块 = "｜" 后面的部分
    if '｜' in biz_val:
        module = biz_val.split('｜')[-1].strip()
    elif '|' in biz_val:
        module = biz_val.split('|')[-1].strip()
    else:
        module = biz_val
    gsec = re.search(r'\*\*Given\*\*(.*?)(?=\*\*|\Z)', b, re.S)
    given = gsec.group(1).strip() if gsec else ''
    wsec = re.search(r'\*\*When\*\*(.*?)(?=\*\*|\Z)', b, re.S)
    when = wsec.group(1).strip() if wsec else ''
    tsec = re.search(r'\*\*Then\*\*(.*?)(?=\*\*依赖\*\*|\*\*弱依赖\*\*|\Z)', b, re.S)
    then = tsec.group(1).strip() if tsec else ''
    entries.append((pid, title, module, coverage, given, when, then))

print(f'parsed {len(entries)} entries')

wb = openpyxl.load_workbook(XLSX)
ws = wb.active
print('sheet:', ws.title, 'max_row:', ws.max_row)

# 确保有"模块"列在用例标识后面(第2列)
hdr = {}
for c in ws[1]:
    if c.value is not None:
        key = str(c.value).strip()
        if key not in hdr:
            hdr[key] = c.column
if '模块' not in hdr:
    ws.insert_cols(2)
    ws.cell(row=1, column=2, value='模块')
    print('inserted 模块 column at B1')
    hdr = {}
    for c in ws[1]:
        if c.value is not None:
            key = str(c.value).strip()
            if key not in hdr:
                hdr[key] = c.column
print('headers:', hdr)

col_id = hdr.get('用例标识')
col_mod = hdr.get('模块')
col_name = hdr.get('用例名称')
col_req = hdr.get('对应需求条目')
col_pre = hdr.get('前置条件')
col_step = hdr.get('测试步骤')
col_exp = hdr.get('预期结果')
for label, col in [('用例标识', col_id), ('模块', col_mod), ('用例名称', col_name),
                   ('对应需求条目', col_req), ('前置条件', col_pre), ('测试步骤', col_step),
                   ('预期结果', col_exp)]:
    assert col, f'缺列: {label}'

# 模板第 2 行是示例行, 保留, 从 max_row+1 追加
row = ws.max_row + 1
empty_given = empty_when = 0
for pid, title, module, coverage, given, when, then in entries:
    ws.cell(row=row, column=col_id, value=pid)
    ws.cell(row=row, column=col_mod, value=module)
    ws.cell(row=row, column=col_name, value=title)
    ws.cell(row=row, column=col_req, value=coverage)
    ws.cell(row=row, column=col_pre, value=given)
    ws.cell(row=row, column=col_step, value=when)
    ws.cell(row=row, column=col_exp, value=then)
    if not given:
        empty_given += 1
    if not when:
        empty_when += 1
    row += 1

wb.save(XLSX)
print(f'written {len(entries)} rows (rows {ws.max_row-len(entries)+1}..{ws.max_row})')
print(f'empty Given: {empty_given}, empty When: {empty_when}')
